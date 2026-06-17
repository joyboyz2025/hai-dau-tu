#!/usr/bin/env python3
"""
Gom transcript cho các video MỚI trong file theo dõi Excel.

CHẠY TRÊN MÁY MAC (nơi có Chrome đã đăng nhập + yt-dlp):
    python3 fetch_new_transcripts.py

- Đọc file Excel "Theo doi video-post kenh dau tu.xlsx" trong cùng thư mục.
- Với mỗi dòng Loại = "Video" và cột Post = "Chưa": tải transcript bằng get_transcript.py
  và lưu JSON vào data/transcripts/<videoId>.json. Bỏ qua video đã có file. Post bỏ qua.

Sau khi chạy xong, lần chạy tự động kế (9h/17h) sẽ tự đọc các file này để tổng hợp & push.
"""
import os, re, sys, subprocess

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Theo doi video-post kenh dau tu.xlsx")
TX_DIR = os.path.join(HERE, "data", "transcripts")
GET = os.path.join(HERE, "get_transcript.py")


def video_id(link: str):
    m = re.search(r"[?&]v=([\w-]{6,})", link or "")
    return m.group(1) if m else None


def main():
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Thiếu openpyxl. Cài: pip3 install openpyxl")
    if not os.path.exists(XLSX):
        sys.exit(f"Không thấy file Excel: {XLSX}")
    os.makedirs(TX_DIR, exist_ok=True)

    wb = load_workbook(XLSX); ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    cL, cLink, cP = idx.get("Loại"), idx.get("Link video/post"), idx.get("Post")
    if None in (cL, cLink, cP):
        sys.exit(f"Không đủ cột. Thấy: {headers}")

    todo = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if (row[cL] or "").strip() == "Video" and (row[cP] or "").strip().lower() in ("chưa","chua",""):
            vid = video_id((row[cLink] or "").strip())
            if vid: todo.append((vid, (row[cLink] or "").strip()))

    if not todo:
        print("Không có video mới cần transcript."); return
    done = skip = fail = 0
    for vid, link in todo:
        out = os.path.join(TX_DIR, f"{vid}.json")
        if os.path.exists(out):
            skip += 1; print(f"• Bỏ qua (đã có): {vid}"); continue
        print(f"• Tải transcript: {vid} ...")
        r = subprocess.run(["python3", GET, link, "--json", "--out", out])
        if r.returncode == 0 and os.path.exists(out):
            done += 1; print(f"  -> OK {out}")
        else:
            fail += 1; print(f"  -> THẤT BẠI {vid} (chưa có phụ đề / cần đăng nhập)")
    print(f"\nXong: {done} mới, {skip} bỏ qua, {fail} thất bại. Transcript ở: {TX_DIR}")


if __name__ == "__main__":
    main()
