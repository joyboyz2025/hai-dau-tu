#!/usr/bin/env python3
"""
Quét video MỚI của từng kênh dựa trên sheet "Kênh" trong file Excel theo dõi,
thêm vào sheet "Video-Post mới" (Post = "Chưa"), rồi tải transcript.

CHẠY TRÊN MÁY MAC (Chrome đã đăng nhập YouTube + yt-dlp cài sẵn):
    python3 scan_channels.py             # quét tất cả kênh "Lên web = Có"
    python3 scan_channels.py --all       # quét cả kênh "Không" (VTV, TC&KD…)
    python3 scan_channels.py --limit 12  # số video mới nhất kiểm mỗi kênh (mặc định 8)
    python3 scan_channels.py --no-transcript   # chỉ thêm dòng, không tải transcript

Nguồn sự thật danh sách kênh: sheet "Kênh" (ID web, Tên kênh khi quét, Link kênh YouTube, Lên web).
Sau khi chạy, các video mới nằm ở sheet "Video-Post mới" với Post = "Chưa" và transcript ở data/transcripts/.
"""
import os, re, sys, json, argparse, subprocess
from datetime import date

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Theo doi video-post kenh dau tu.xlsx")
TX_DIR = os.path.join(HERE, "data", "transcripts")
GET = os.path.join(HERE, "get_transcript.py")


def video_id(link: str):
    m = re.search(r"[?&]v=([\w-]{6,})", link or "")
    if m:
        return m.group(1)
    m = re.search(r"youtu\.be/([\w-]{6,})", link or "")
    return m.group(1) if m else None


def list_channel_videos(channel_url: str, limit: int):
    """Trả về list dict {id,title,url,date} cho `limit` video mới nhất của kênh."""
    url = channel_url.rstrip("/")
    if not url.endswith("/videos"):
        url += "/videos"
    cmd = [
        "yt-dlp", "--cookies-from-browser", "chrome",
        "--flat-playlist", "--playlist-end", str(limit),
        "--print", "%(id)s\t%(title)s\t%(webpage_url)s\t%(upload_date)s",
        url,
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    out = []
    for line in r.stdout.splitlines():
        parts = line.split("\t")
        if len(parts) >= 3 and parts[0]:
            vid, title, vurl = parts[0], parts[1], parts[2]
            up = parts[3] if len(parts) > 3 and parts[3] not in ("NA", "") else ""
            date_str = f"{up[6:8]}/{up[4:6]}/{up[0:4]}" if len(up) == 8 else ""
            out.append({"id": vid, "title": title, "url": vurl, "date": date_str})
    if not out and r.returncode != 0:
        sys.stderr.write(f"  [cảnh báo] không quét được: {url}\n  {r.stderr.strip()[:200]}\n")
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--all", action="store_true", help="quét cả kênh Lên web = Không")
    ap.add_argument("--limit", type=int, default=8)
    ap.add_argument("--no-transcript", action="store_true")
    args = ap.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Thiếu openpyxl. Cài: pip3 install openpyxl")
    if not os.path.exists(XLSX):
        sys.exit(f"Không thấy file Excel: {XLSX}")

    wb = load_workbook(XLSX)
    if "Kênh" not in wb.sheetnames:
        sys.exit('Thiếu sheet "Kênh" trong file Excel.')
    ks = wb["Kênh"]
    kh = {ks.cell(row=1, column=c).value: c for c in range(1, ks.max_column + 1)}
    need = ["Tên kênh (khi quét)", "Link kênh YouTube", "Lên web"]
    if any(k not in kh for k in need):
        sys.exit(f"Sheet Kênh thiếu cột. Thấy: {list(kh)}")

    channels = []
    for r in range(2, ks.max_row + 1):
        link = (ks.cell(row=r, column=kh["Link kênh YouTube"]).value or "").strip()
        name = (ks.cell(row=r, column=kh["Tên kênh (khi quét)"]).value or "").strip()
        web = (ks.cell(row=r, column=kh["Lên web"]).value or "").strip()
        if not link or not name:
            continue
        if not args.all and web.lower() not in ("có", "co", "yes", "x"):
            continue
        channels.append((name, link))

    if not channels:
        sys.exit("Không có kênh nào có Link YouTube (điền cột 'Link kênh YouTube' ở sheet Kênh).")

    vp = wb["Video-Post mới"]
    vph = {vp.cell(row=1, column=c).value: c for c in range(1, vp.max_column + 1)}
    cLink = vph["Link video/post"]
    existing = set()
    for r in range(2, vp.max_row + 1):
        vid = video_id(str(vp.cell(row=r, column=cLink).value or ""))
        if vid:
            existing.add(vid)

    today = date.today().strftime("%d/%m/%Y")
    added = []
    for name, link in channels:
        print(f"• Quét kênh: {name}")
        vids = list_channel_videos(link, args.limit)
        for v in vids:
            if v["id"] in existing:
                continue
            existing.add(v["id"])
            vp.append([name, "Video", v["title"], v["url"], v["date"], today, "Chưa"])
            added.append(v)
            print(f"    + MỚI: {v['id']}  {v['title'][:60]}")

    wb.save(XLSX)
    print(f"\nĐã thêm {len(added)} video mới vào 'Video-Post mới'.")

    if args.no_transcript or not added:
        return
    os.makedirs(TX_DIR, exist_ok=True)
    done = fail = 0
    for v in added:
        out = os.path.join(TX_DIR, f"{v['id']}.json")
        if os.path.exists(out):
            continue
        print(f"• Tải transcript: {v['id']} …")
        r = subprocess.run(["python3", GET, v["url"], "--json", "--out", out])
        if r.returncode == 0 and os.path.exists(out):
            done += 1
        else:
            fail += 1
            print(f"  -> THẤT BẠI {v['id']} (có thể là hội viên cấp cao / chưa có phụ đề)")
    print(f"Transcript: {done} tải mới, {fail} thất bại. Ở: {TX_DIR}")


if __name__ == "__main__":
    main()
